import csv
from datetime import datetime, timedelta
from openpyxl import load_workbook

def combine_and_export_data_fixed(input_file):
    # Load the workbook
    wb = load_workbook(input_file)

    # Access the correct sheets by name or index
    elec_sheet = wb[wb.sheetnames[1]]  # Second sheet - Electricity kWh
    weather_sheet = wb[wb.sheetnames[2]]  # Third sheet - Weather Archive

    # Read headers (assuming headers are on row 2)
    elec_headers = [cell.value for cell in elec_sheet[2]]
    weather_headers = [cell.value for cell in weather_sheet[2]]

    # Locate indices
    elec_time_idx = elec_headers.index('Timestamps')
    weather_time_idx = weather_headers.index('Timestamps')

    # Desired weather fields
    desired_weather_headers = ['air temperature', 'Atm pressure mm of mercury', 'Relative humidity (%)']
    weather_indices = [weather_headers.index(h) for h in desired_weather_headers]

    # Extract electricity data
    elec_data = []
    for row in elec_sheet.iter_rows(min_row=3, values_only=True):
        ts = row[elec_time_idx]
        if isinstance(ts, datetime):
            elec_data.append({
                'timestamp': ts,
                'values': row[1:11]  # Columns B to K (Ap1 to Ap10)
            })

    # Extract weather data
    weather_data = []
    for row in weather_sheet.iter_rows(min_row=3, values_only=True):
        ts = row[weather_time_idx]
        if isinstance(ts, datetime):
            values = [row[idx] for idx in weather_indices]
            weather_data.append({'timestamp': ts, 'values': values})

    # Match each electricity record with the closest weather record within 30 minutes
    combined = []
    for elec in elec_data:
        best_match = None
        min_diff = timedelta(minutes=30)

        for weather in weather_data:
            diff = abs(weather['timestamp'] - elec['timestamp'])
            if diff <= min_diff:
                min_diff = diff
                best_match = weather

        if best_match:
            combined.append({
                'timestamp': elec['timestamp'],
                'elec_values': elec['values'],
                'weather_values': best_match['values']
            })

    # Write output to a single CSV file
    output_filename = 'data/outputs/combined_data.csv'
    with open(output_filename, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['Timestamps'] + elec_headers[1:11] + desired_weather_headers)
        for record in combined:
            row = [record['timestamp'].strftime('%Y-%m-%d %H:%M:%S')] + \
                  list(record['elec_values']) + \
                  list(record['weather_values'])
            writer.writerow(row)

    print(f"Created: {output_filename}")

# Example usage
input_file = 'data/Buildings_aligned.xlsx'
combine_and_export_data_fixed(input_file)
